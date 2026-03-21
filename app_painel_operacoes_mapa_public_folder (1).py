from __future__ import annotations

import html
import re
import tempfile
import urllib.parse
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl import load_workbook


# ============================================================
# CONFIGURAÇÃO FIXA DA PASTA PÚBLICA DO GOOGLE DRIVE
# ============================================================
PUBLIC_FOLDER_URL = "https://drive.google.com/drive/folders/1zEHRpVyvHQ8PQve2RWhJRYqgVpjvymUf?usp=sharing"
PUBLIC_FOLDER_ID = "1zEHRpVyvHQ8PQve2RWhJRYqgVpjvymUf"
PUBLIC_FILENAME_CONTAINS = "Pipeline"
EXTENSOES_VALIDAS = {".xlsx", ".xlsm", ".xltx", ".xltm"}


# ============================================================
# IDENTIDADE VISUAL
# ============================================================
COLOR_BG = "#F3F7FA"
COLOR_CARD = "#FFFFFF"
COLOR_BORDER = "#D9E6F0"
COLOR_TEXT = "#15384F"
COLOR_MUTED = "#5E768B"
COLOR_DARK = "#0D4868"
COLOR_DARK_2 = "#123A5A"
COLOR_LIGHT = "#D9EDF7"
COLOR_LIGHT_2 = "#EDF7FC"
COLOR_ACCENT = "#2BA6C1"
COLOR_WHITE = "#FFFFFF"


# ============================================================
# CONFIG STREAMLIT
# ============================================================
st.set_page_config(
    page_title="Painel de Operações | MAPA",
    layout="wide",
    initial_sidebar_state="collapsed",
)


# ============================================================
# ESTILO GLOBAL
# ============================================================
GLOBAL_CSS = f"""
<style>
    .stApp {{
        background: {COLOR_BG};
        color: {COLOR_TEXT};
    }}

    div[data-testid="stMetric"] {{
        background: {COLOR_CARD};
        border: 1px solid {COLOR_BORDER};
        border-radius: 14px;
        padding: 10px 12px;
        box-shadow: 0 2px 10px rgba(13, 72, 104, 0.05);
    }}

    .mapa-title {{
        font-size: 2rem;
        font-weight: 800;
        color: {COLOR_DARK_2};
        margin: 0 0 2px 0;
        letter-spacing: -0.02em;
    }}

    .mapa-subtitle {{
        color: {COLOR_MUTED};
        margin-bottom: 18px;
        font-size: 0.98rem;
    }}

    .mapa-section-title {{
        font-size: 1.1rem;
        font-weight: 700;
        color: {COLOR_DARK_2};
        margin: 14px 0 10px 0;
    }}

    .mapa-chip {{
        display: inline-block;
        padding: 5px 10px;
        border-radius: 999px;
        background: {COLOR_LIGHT_2};
        border: 1px solid {COLOR_BORDER};
        color: {COLOR_DARK_2};
        font-size: 0.85rem;
        margin-right: 8px;
    }}

    .table-wrap {{
        overflow-x: auto;
        border: 1px solid {COLOR_BORDER};
        border-radius: 16px;
        background: {COLOR_CARD};
        box-shadow: 0 2px 10px rgba(13, 72, 104, 0.05);
    }}

    .table-inner {{
        min-width: 1780px;
    }}

    .row-grid {{
        display: grid;
        grid-template-columns:
            56px 82px 2.0fr 1.35fr 0.95fr 1.1fr 0.95fr 1fr 1.1fr 1.15fr 1fr 1fr 1.2fr 1.2fr;
        align-items: stretch;
    }}

    .header-row {{
        background: {COLOR_DARK_2};
        color: {COLOR_WHITE};
        border-bottom: 1px solid rgba(255,255,255,0.08);
        position: sticky;
        top: 0;
        z-index: 10;
    }}

    .header-cell {{
        padding: 12px 10px;
        font-size: 0.84rem;
        font-weight: 700;
        border-right: 1px solid rgba(255,255,255,0.08);
        white-space: nowrap;
    }}

    .header-cell:last-child,
    .body-cell:last-child {{
        border-right: none;
    }}

    .header-link {{
        color: {COLOR_WHITE};
        text-decoration: none;
        display: block;
    }}

    .header-link:hover {{
        text-decoration: underline;
    }}

    .body-row {{
        background: {COLOR_CARD};
        border-bottom: 1px solid {COLOR_BORDER};
    }}

    .body-row:hover {{
        background: #F8FBFD;
    }}

    .body-row.active {{
        background: {COLOR_DARK};
    }}

    .body-cell {{
        padding: 12px 10px;
        font-size: 0.86rem;
        color: {COLOR_TEXT};
        border-right: 1px solid {COLOR_BORDER};
        display: flex;
        align-items: center;
        min-height: 62px;
    }}

    .body-row.active .body-cell {{
        color: {COLOR_WHITE};
        border-right: 1px solid rgba(255,255,255,0.08);
    }}

    .cell-link {{
        color: inherit;
        text-decoration: none;
        display: block;
        width: 100%;
        white-space: normal;
        word-break: break-word;
    }}

    .cell-link:hover {{
        text-decoration: underline;
    }}

    .detail-box {{
        background: {COLOR_LIGHT};
        border-top: 1px solid rgba(13, 72, 104, 0.12);
        border-bottom: 1px solid rgba(13, 72, 104, 0.12);
        padding: 0;
    }}

    .detail-line {{
        padding: 11px 14px;
        border-bottom: 1px solid rgba(13, 72, 104, 0.10);
        color: {COLOR_DARK_2};
        font-size: 0.9rem;
        white-space: pre-wrap;
        word-break: break-word;
    }}

    .detail-line:last-child {{
        border-bottom: none;
    }}

    .detail-label {{
        font-style: italic;
        font-weight: 700;
        display: block;
        margin-bottom: 2px;
    }}

    .empty-state {{
        padding: 22px 14px;
        color: {COLOR_MUTED};
    }}
</style>
"""

st.markdown(GLOBAL_CSS, unsafe_allow_html=True)


# ============================================================
# UTILITÁRIOS
# ============================================================
def format_brl(value: Any) -> str:
    try:
        number = float(value or 0)
        return f"R$ {number:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "R$ 0,00"



def format_datetime(value: Any) -> str:
    if value is None or pd.isna(value):
        return "—"
    try:
        return pd.to_datetime(value).strftime("%d/%m/%Y")
    except Exception:
        return str(value)



def normalize_yes_no(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    mapping = {
        "sim": "Sim",
        "não": "Não",
        "nao": "Não",
        "n/a": "N/A",
        "na": "NA",
    }
    return mapping.get(text, str(value).strip())



def to_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    text = text.replace("R$", "").replace("%", "").replace(" ", "")

    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")

    try:
        return float(text)
    except Exception:
        return None



def extrair_probabilidade(chance: Any) -> float:
    if chance is None:
        return 0.0
    text = str(chance).strip().lower()
    if text.startswith("1") or "alta" in text:
        return 0.75
    if text.startswith("2") or "média" in text or "media" in text:
        return 0.50
    if text.startswith("3") or "baixa" in text:
        return 0.25
    return 0.0



def extract_folder_id(value: Optional[str]) -> str:
    if not value:
        return ""
    text = str(value).strip()
    match = re.search(r"/folders/([a-zA-Z0-9_-]+)", text)
    if match:
        return match.group(1)
    if re.fullmatch(r"[a-zA-Z0-9_-]{20,}", text):
        return text
    return ""



def infer_sort_key_from_name(path: Path) -> tuple:
    name = path.name.lower()
    date_match = re.search(r"(20\d{{2}}[-_]?\d{{2}}[-_]?\d{{2}})", name)
    if date_match:
        raw = re.sub(r"[^0-9]", "", date_match.group(1))
        return (raw, name)
    return ("", name)



def html_escape(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "—"
    text = str(value).strip()
    if not text:
        return "—"
    return html.escape(text)



def text_to_html_block(value: Any) -> str:
    text = html_escape(value)
    return text.replace("\n", "<br>")



def safe_secret_get(section: str, key: str, default: str = "") -> str:
    try:
        if section in st.secrets and key in st.secrets[section]:
            return str(st.secrets[section][key]).strip()
    except Exception:
        pass
    return default



def get_public_drive_config() -> Dict[str, str]:
    folder_url = safe_secret_get("public_google_drive", "folder_url", PUBLIC_FOLDER_URL)
    folder_id = safe_secret_get("public_google_drive", "folder_id", PUBLIC_FOLDER_ID)
    filename_contains = safe_secret_get(
        "public_google_drive", "filename_contains", PUBLIC_FILENAME_CONTAINS
    )

    return {
        "folder_url": folder_url or PUBLIC_FOLDER_URL,
        "folder_id": extract_folder_id(folder_id) or extract_folder_id(folder_url) or PUBLIC_FOLDER_ID,
        "filename_contains": filename_contains or PUBLIC_FILENAME_CONTAINS,
    }


# ============================================================
# LEITURA DA PASTA PÚBLICA DO GOOGLE DRIVE
# ============================================================
@st.cache_data(show_spinner=False, ttl=600)
def list_public_drive_excel_files(
    folder_url: Optional[str] = None,
    folder_id: Optional[str] = None,
    filename_contains: Optional[str] = None,
) -> List[str]:
    import gdown

    resolved_url = (folder_url or "").strip() or PUBLIC_FOLDER_URL
    resolved_id = extract_folder_id(folder_id) or extract_folder_id(resolved_url) or PUBLIC_FOLDER_ID
    name_filter = (filename_contains or "").strip() or PUBLIC_FILENAME_CONTAINS

    if not resolved_id and not resolved_url:
        raise ValueError("A pasta pública do Google Drive não foi configurada.")

    cache_dir = Path(tempfile.gettempdir()) / "mapa_painel_public_folder"
    cache_dir.mkdir(parents=True, exist_ok=True)

    downloaded: Optional[Iterable[str]] = None
    last_error: Optional[Exception] = None

    download_attempts = [
        {"id": resolved_id, "url": None},
        {"id": None, "url": resolved_url},
    ]

    for attempt in download_attempts:
        try:
            kwargs = {
                "quiet": True,
                "remaining_ok": True,
                "use_cookies": False,
            }
            if attempt["id"]:
                kwargs["id"] = attempt["id"]
            if attempt["url"]:
                kwargs["url"] = attempt["url"]

            try:
                kwargs["output"] = str(cache_dir)
                downloaded = gdown.download_folder(**kwargs)
            except TypeError:
                kwargs.pop("output", None)
                downloaded = gdown.download_folder(**kwargs)

            if downloaded:
                break
        except Exception as exc:  # pragma: no cover - comportamento externo
            last_error = exc

    if not downloaded:
        message = (
            "Não foi possível listar os arquivos da pasta pública do Google Drive. "
            "Confira se a pasta está em 'Qualquer pessoa com o link'."
        )
        if last_error:
            message += f" Detalhe: {last_error}"
        raise FileNotFoundError(message)

    paths: List[Path] = []
    for item in downloaded:
        try:
            path = Path(item)
            if path.is_file() and path.suffix.lower() in EXTENSOES_VALIDAS:
                if not name_filter or name_filter.lower() in path.name.lower():
                    paths.append(path)
        except Exception:
            continue

    unique_paths = list({str(p.resolve()): p for p in paths}.values())
    unique_paths.sort(key=lambda p: infer_sort_key_from_name(p), reverse=True)

    return [str(p) for p in unique_paths]


# ============================================================
# PARSER DO EXCEL
# ============================================================
LABEL_MAP = {
    "Cliente:": "cliente",
    "Top Five": "top_five",
    "Operação:": "operacao",
    "Prioridade:": "prioridade",
    "Status:": "status",
    "Responsável:": "responsavel",
    "Chance de Fechamento:": "chance_fechamento",
    "Resumo:": "resumo",
    "Valor da Operação:": "valor_operacao",
    "Forma de Remuneração": "forma_remuneracao",
    "Comissão Total:": "comissao_total",
    "Comisão Mapa:": "comissao_mapa",
    "Comissão Mapa:": "comissao_mapa",
    "Aprovação no Comitê?": "aprovacao_comite",
    "Mandato?": "mandato",
}


@st.cache_data(show_spinner=False, ttl=600)
def parse_pipeline_excel_from_path(file_path: str) -> pd.DataFrame:
    workbook = load_workbook(filename=file_path, data_only=True)
    if "Pipeline" not in workbook.sheetnames:
        raise ValueError("A planilha selecionada não possui a aba 'Pipeline'.")

    ws = workbook["Pipeline"]
    records: List[Dict[str, Any]] = []
    current: Optional[Dict[str, Any]] = None

    for row in range(1, ws.max_row + 1):
        index_value = ws.cell(row=row, column=1).value
        col_b = ws.cell(row=row, column=2).value
        col_d = ws.cell(row=row, column=4).value
        col_e = ws.cell(row=row, column=5).value
        col_g = ws.cell(row=row, column=7).value

        if col_b == "Cliente:":
            if current:
                records.append(current)

            current = {
                "id": index_value,
                "cliente": col_d,
                "atualizacao": col_g,
                "top_five": None,
                "operacao": None,
                "prioridade": None,
                "status": None,
                "responsavel": None,
                "chance_fechamento": None,
                "resumo": None,
                "valor_operacao": None,
                "forma_remuneracao": None,
                "comissao_total": None,
                "comissao_mapa": None,
                "aprovacao_comite": None,
                "mandato": None,
                "status_detalhado": None,
                "link_apresentacao": None,
                "link_documentos": None,
                "historicos_disponiveis": None,
            }
            continue

        if current is None:
            continue

        if col_b in LABEL_MAP:
            field_name = LABEL_MAP[col_b]
            current[field_name] = col_d
            if col_b == "Operação:" and col_e not in (None, ""):
                current["status_detalhado"] = col_e

    if current:
        records.append(current)

    df = pd.DataFrame(records)
    if df.empty:
        return df

    min_fields = ["cliente", "operacao", "status", "responsavel", "valor_operacao"]
    df = df[df[min_fields].notna().any(axis=1)].copy()

    for col in ["valor_operacao", "comissao_total", "comissao_mapa"]:
        df[col] = df[col].apply(to_float).fillna(0.0)

    df["atualizacao"] = pd.to_datetime(df["atualizacao"], errors="coerce")
    df["top_five"] = df["top_five"].apply(normalize_yes_no)
    df["prob_fechamento"] = df["chance_fechamento"].apply(extrair_probabilidade)
    df["valor_ponderado"] = df["valor_operacao"].fillna(0.0) * df["prob_fechamento"]
    id_series = pd.to_numeric(df["id"], errors="coerce")
    fallback_ids = pd.Series(range(1, len(df) + 1), index=df.index, dtype="float64")
    df["id"] = id_series.where(id_series.notna(), fallback_ids).astype(int)

    def compose_status(row: pd.Series) -> str:
        parts: List[str] = []
        if row.get("status_detalhado"):
            parts.append(str(row.get("status_detalhado")).strip())
        if row.get("resumo"):
            parts.append(str(row.get("resumo")).strip())
        return "\n\n".join([p for p in parts if p]) or "—"

    df["status_expandido"] = df.apply(compose_status, axis=1)
    df["historicos_disponiveis"] = df["historicos_disponiveis"].fillna("—")
    df["link_apresentacao"] = df["link_apresentacao"].fillna("—")
    df["link_documentos"] = df["link_documentos"].fillna("—")

    return df.sort_values("id").reset_index(drop=True)


# ============================================================
# SORT / QUERY PARAMS
# ============================================================
COLUMN_SPECS = [
    ("id", "#"),
    ("top_five", "Top Five"),
    ("cliente", "Cliente"),
    ("operacao", "Operação"),
    ("prioridade", "Prioridade"),
    ("status", "Status"),
    ("responsavel", "Responsável"),
    ("chance_fechamento", "Chance de Fechamento"),
    ("valor_operacao", "Valor da Operação"),
    ("forma_remuneracao", "Forma de Remuneração"),
    ("comissao_total", "Comissão Total"),
    ("comissao_mapa", "Comissão Mapa"),
    ("aprovacao_comite", "Aprovação no Comitê?"),
    ("mandato", "Mandato?"),
]

SORTABLE_COLUMNS = {key for key, _ in COLUMN_SPECS}
DEFAULT_SORT = "id"
DEFAULT_DIRECTION = "asc"



def get_query_value(name: str, default: str = "") -> str:
    try:
        value = st.query_params.get(name, default)
        if isinstance(value, list):
            return str(value[-1]) if value else default
        return str(value)
    except Exception:
        return default



def get_current_query_params() -> Dict[str, str]:
    current: Dict[str, str] = {}
    try:
        for key, value in st.query_params.items():
            if isinstance(value, list):
                current[key] = str(value[-1]) if value else ""
            else:
                current[key] = str(value)
    except Exception:
        pass
    return current



def build_query_string(current: Dict[str, str], **updates: Optional[str]) -> str:
    merged = dict(current)
    for key, value in updates.items():
        if value in (None, ""):
            merged.pop(key, None)
        else:
            merged[key] = str(value)
    return "?" + urllib.parse.urlencode(merged)



def apply_sort(df: pd.DataFrame, sort_by: str, direction: str) -> pd.DataFrame:
    if sort_by not in df.columns:
        sort_by = DEFAULT_SORT

    ascending = direction != "desc"
    sorted_df = df.copy()

    if sort_by in {"valor_operacao", "valor_ponderado", "comissao_total", "comissao_mapa", "id"}:
        sorted_df[sort_by] = pd.to_numeric(sorted_df[sort_by], errors="coerce")
    elif sort_by == "atualizacao":
        sorted_df[sort_by] = pd.to_datetime(sorted_df[sort_by], errors="coerce")
    else:
        sorted_df[sort_by] = sorted_df[sort_by].fillna("").astype(str).str.lower()

    return sorted_df.sort_values(by=sort_by, ascending=ascending, na_position="last").reset_index(drop=True)



def format_cell(column: str, value: Any) -> str:
    if column in {"valor_operacao", "valor_ponderado", "comissao_total", "comissao_mapa"}:
        return format_brl(value)
    if column == "atualizacao":
        return format_datetime(value)
    return html_escape(value)


# ============================================================
# RENDER DA TABELA EXPANSÍVEL
# ============================================================

def build_expandable_table_html(df: pd.DataFrame, sort_by: str, direction: str, open_row_id: str) -> str:
    current_params = get_current_query_params()
    pieces: List[str] = ["<div class='table-wrap'><div class='table-inner'>"]

    # Cabeçalho
    pieces.append("<div class='row-grid header-row'>")
    for key, label in COLUMN_SPECS:
        next_direction = "asc"
        arrow = ""
        if sort_by == key:
            if direction == "asc":
                next_direction = "desc"
                arrow = " ▲"
            else:
                next_direction = "asc"
                arrow = " ▼"
        href = build_query_string(current_params, sort=key, dir=next_direction, open=None)
        pieces.append(
            f"<div class='header-cell'><a class='header-link' href='{href}'>{html.escape(label + arrow)}</a></div>"
        )
    pieces.append("</div>")

    # Corpo
    for _, row in df.iterrows():
        row_id = str(int(row["id"]))
        is_open = open_row_id == row_id
        row_href = build_query_string(current_params, open=None if is_open else row_id)
        row_class = "body-row active" if is_open else "body-row"

        pieces.append(f"<div class='row-grid {row_class}'>")
        for key, _ in COLUMN_SPECS:
            pieces.append(
                "<div class='body-cell'>"
                f"<a class='cell-link' href='{row_href}'>{format_cell(key, row.get(key))}</a>"
                "</div>"
            )
        pieces.append("</div>")

        if is_open:
            pieces.append("<div class='detail-box'>")
            detail_lines = [
                ("Link para Apresentação:", row.get("link_apresentacao") or "—"),
                ("Link para Acesso aos Documentos:", row.get("link_documentos") or "—"),
                ("Status:", row.get("status_expandido") or "—"),
                ("Históricos Disponíveis:", row.get("historicos_disponiveis") or "—"),
            ]
            for label, value in detail_lines:
                pieces.append(
                    "<div class='detail-line'>"
                    f"<span class='detail-label'>{html.escape(label)}</span>"
                    f"{text_to_html_block(value)}"
                    "</div>"
                )
            pieces.append("</div>")

    if df.empty:
        pieces.append("<div class='empty-state'>Nenhuma operação encontrada com os filtros aplicados.</div>")

    pieces.append("</div></div>")
    return "".join(pieces)


# ============================================================
# APP
# ============================================================

def main() -> None:
    st.markdown("<div class='mapa-title'>Painel de Operações</div>", unsafe_allow_html=True)
    st.markdown(
        "<div class='mapa-subtitle'>Leitura automática da pasta pública do Google Drive com card expansível por operação.</div>",
        unsafe_allow_html=True,
    )

    config = get_public_drive_config()
    st.markdown(
        (
            f"<span class='mapa-chip'>Fonte: pasta pública do Google Drive</span>"
            f"<span class='mapa-chip'>Filtro de nome: {html.escape(config['filename_contains'])}</span>"
        ),
        unsafe_allow_html=True,
    )

    with st.spinner("Carregando arquivos da pasta pública..."):
        files = list_public_drive_excel_files(
            folder_url=config["folder_url"],
            folder_id=config["folder_id"],
            filename_contains=config["filename_contains"],
        )

    if not files:
        st.error(
            "Nenhum arquivo Excel foi encontrado na pasta pública do Google Drive. "
            "Verifique o compartilhamento e o filtro de nome."
        )
        st.stop()

    file_options = {Path(f).name: f for f in files}
    selected_name = st.selectbox("Planilha", options=list(file_options.keys()), index=0)
    selected_path = file_options[selected_name]

    df = parse_pipeline_excel_from_path(selected_path)
    if df.empty:
        st.warning("Nenhuma operação válida foi encontrada na aba 'Pipeline'.")
        st.stop()

    st.markdown("<div class='mapa-section-title'>Filtros</div>", unsafe_allow_html=True)
    col1, col2, col3, col4, col5, col6 = st.columns([1.2, 1.2, 1.2, 1.1, 1.1, 1.5])

    responsaveis = sorted([x for x in df["responsavel"].dropna().unique().tolist() if str(x).strip()])
    statuses = sorted([x for x in df["status"].dropna().unique().tolist() if str(x).strip()])
    operacoes = sorted([x for x in df["operacao"].dropna().unique().tolist() if str(x).strip()])
    prioridades = sorted([x for x in df["prioridade"].dropna().unique().tolist() if str(x).strip()])

    with col1:
        filtro_responsavel = st.multiselect("Responsável", responsaveis, default=responsaveis)
    with col2:
        filtro_status = st.multiselect("Status", statuses, default=statuses)
    with col3:
        filtro_operacao = st.multiselect("Operação", operacoes, default=operacoes)
    with col4:
        filtro_prioridade = st.multiselect("Prioridade", prioridades, default=prioridades)
    with col5:
        filtro_top_five = st.selectbox(
            "Top Five",
            options=["Todas", "Somente Top Five", "Somente Não Top Five"],
            index=0,
        )
    with col6:
        filtro_busca = st.text_input("Buscar cliente", placeholder="Digite parte do nome...")

    filtered = df.copy()
    if filtro_responsavel:
        filtered = filtered[filtered["responsavel"].isin(filtro_responsavel)]
    if filtro_status:
        filtered = filtered[filtered["status"].isin(filtro_status)]
    if filtro_operacao:
        filtered = filtered[filtered["operacao"].isin(filtro_operacao)]
    if filtro_prioridade:
        filtered = filtered[filtered["prioridade"].isin(filtro_prioridade)]
    if filtro_top_five == "Somente Top Five":
        filtered = filtered[filtered["top_five"] == "Sim"]
    elif filtro_top_five == "Somente Não Top Five":
        filtered = filtered[filtered["top_five"] != "Sim"]
    if filtro_busca:
        filtered = filtered[
            filtered["cliente"].fillna("").astype(str).str.contains(filtro_busca, case=False, na=False)
        ]

    total_operacoes = int(len(filtered))
    valor_total = float(filtered["valor_operacao"].sum())
    valor_ponderado = float(filtered["valor_ponderado"].sum())
    comissao_mapa_total = float(filtered["comissao_mapa"].sum())
    ticket_medio = float(filtered["valor_operacao"].mean()) if total_operacoes else 0.0

    st.markdown("<div class='mapa-section-title'>Métricas</div>", unsafe_allow_html=True)
    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Nº Operações", f"{total_operacoes}")
    m2.metric("Valor Total", format_brl(valor_total))
    m3.metric("Valor Ponderado", format_brl(valor_ponderado))
    m4.metric("Comissão MAPA", format_brl(comissao_mapa_total))
    m5.metric("Ticket Médio", format_brl(ticket_medio))

    chart_col1, chart_col2 = st.columns(2)
    with chart_col1:
        base_status = (
            filtered.groupby("status", dropna=False)["valor_operacao"]
            .sum()
            .reset_index()
            .sort_values("valor_operacao", ascending=False)
        )
        fig_status = px.bar(base_status, x="status", y="valor_operacao", title="Valor por Status")
        fig_status.update_layout(height=340, xaxis_title="", yaxis_title="Valor da Operação")
        st.plotly_chart(fig_status, use_container_width=True)

    with chart_col2:
        base_oper = (
            filtered.groupby("operacao", dropna=False)["valor_operacao"]
            .sum()
            .reset_index()
            .sort_values("valor_operacao", ascending=False)
        )
        fig_oper = px.bar(base_oper, x="operacao", y="valor_operacao", title="Valor por Operação")
        fig_oper.update_layout(height=340, xaxis_title="", yaxis_title="Valor da Operação")
        st.plotly_chart(fig_oper, use_container_width=True)

    st.markdown("<div class='mapa-section-title'>Lista de Operações</div>", unsafe_allow_html=True)

    sort_by = get_query_value("sort", DEFAULT_SORT)
    direction = get_query_value("dir", DEFAULT_DIRECTION)
    open_row_id = get_query_value("open", "")

    sorted_df = apply_sort(filtered, sort_by if sort_by in SORTABLE_COLUMNS else DEFAULT_SORT, direction)

    table_html = build_expandable_table_html(
        sorted_df,
        sort_by if sort_by in SORTABLE_COLUMNS else DEFAULT_SORT,
        direction if direction in {"asc", "desc"} else DEFAULT_DIRECTION,
        open_row_id,
    )
    st.markdown(table_html, unsafe_allow_html=True)

    csv_export = filtered.to_csv(index=False).encode("utf-8-sig")
    st.download_button(
        label="Baixar CSV filtrado",
        data=csv_export,
        file_name="painel_operacoes_filtrado.csv",
        mime="text/csv",
    )


if __name__ == "__main__":
    main()
